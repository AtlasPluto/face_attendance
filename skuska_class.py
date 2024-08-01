import cv2
from deepface import DeepFace
import os
import customtkinter as cTk
from datetime import datetime as dt
from datetime import timedelta
import openpyxl
from openpyxl.styles import Border, Side, colors
from PIL import Image, ImageTk


cTk.set_appearance_mode("system")  # Modes: system (default), light, dark
cTk.set_default_color_theme("green")  # Themes: blue (default), dark-blue, green

gui = cTk.CTk()  # vytvori graficku plochu
gui.resizable(False, False)
gui.geometry("850x510")


class FaceFunctions:
    def __init__(self) -> None:
        self.video_capture = cv2.VideoCapture(0)
        self.face_classifier = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
        
    def detect_bounding_box(self):
        if len(self.frame.shape) == 3 and self.frame.shape[2] == 3:
            # zmen frame na sedy colorspace
            gray_image = cv2.cvtColor(self.frame, cv2.COLOR_BGR2GRAY)
        else:
            # ak je sedi nie je treba ho menit
            gray_image = self.frame
        faces = self.face_classifier.detectMultiScale(gray_image, 1.1, 5, minSize=(40, 40))
        for (x, y, w, h) in faces:  
            current_face = self.frame[y:y+h, x:x+w]   # bereme tvar po tvari, ak viacej ludi tak aby sa nezobralo viacej na raz
            identita = self.over(current_face,"DATA/TVARE")  # ziskanie identity
            # print(identita)
            excel.uloz_dochadzku(identita) # ulozenie dochadzky
            
            
            # vykreslenie stvorca okolo tvare aj mena ale to pojde prec
            cv2.rectangle(self.frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
            cv2.rectangle(self.frame, (x, y +h), (x + w, y + h + 30), (0, 255, 0), 2)
            cv2.putText(self.frame,identita, (x,y+h + 22),cv2.FONT_ITALIC,1, (255,255,255), 1, cv2.LINE_AA )

        return faces
    
    
    def over(self,img_path1,database_path):
        try :
            vysledok = DeepFace.find(img_path1,database_path,model_name="VGG-Face", enforce_detection=False,silent=True)
            identita = vysledok[0]['identity']
            if  not identita.empty and identita[0] != "nepoznam" :
                file_paths = identita[0]
                filenames = [os.path.splitext(os.path.basename(file_path))[0] for file_path in file_paths]
                file_png = filenames[len("DATA/TVARE/"):]
                file_string = ''.join(file_png)
                last_dot = file_string.rindex('.')
                filename = file_string[0:last_dot]
                return filename
            else:
                return "nepoznam"
        except ValueError :
            return f'nepoznam'
        
    def ziskaj_tvar(self,faces,meno):
        if len(faces) > 0:
            #print("halo halo halo")
            (x, y, w, h) = faces[0]
            face = self.frame[y:y + h, x:x + w]
            frame_face = cv2.resize(face,(300,300))

            otazka = self.over(frame_face,"DATA/TVARE")
            if otazka == "nepoznam" :
                img_path = f"DATA/TVARE/{meno}.png"
                cv2.imwrite(img_path,frame_face)
                excel.ulozenie_mena(meno)
                #cv2.destroyWindow(f'{c_tvare}')
                if os.path.exists(f'DATA/TVARE/representations_vgg_face.pkl') :
                    os.remove(f'DATA/TVARE/representations_vgg_face.pkl')
                else :
                    pass
            else:
                pass
                #print("uz tam je")
                
                
class CustomTkinter : 
    def __init__(self,g,m,tu,w,th,f,people) -> None:
        #gui
        self.gui = g
        self.off_img_path = "IMG/off2.png"  
        self.off_image_pil = Image.open(self.off_img_path)
        self.off_img = ImageTk.PhotoImage(self.off_image_pil)
        # zoznamy pre nacitanie veci 
        self.mondays = m
        self.tuesdays = tu
        self.wednesday = w
        self.thursday = th
        self.frydays = f
        self.all_the_people = people
        self.dni_str = excel.datum_tyzdna()
        # main values (kamera a main frame)
        self.frame_kamera = cTk.CTkFrame(master=self.gui, width=680, height=520, border_width=2, border_color="green")
        self.frame_kamera.place(relx=1, rely=0, anchor="ne")
        self.label_kamera = cTk.CTkLabel(self.frame_kamera, anchor='center', text="", width=640, height=480)
        self.label_kamera.pack(pady=15, padx=15)
        # slider values
        self.KAMERA_SLIDER_VALUE = 0
        self.DETEKCIA_SLIDER_VALUE = 0
        # main screen frame , 
        self.frame_control = cTk.CTkFrame(master=self.gui, width=75, height=700, border_width=2, border_color="green")
        self.frame_control.place(relx=0, rely=0, anchor="nw",relheight=1.0)
        ####################################################################
        self.MENU_butn = cTk.CTkButton(master= self.frame_control, text="MENU",width=120,fg_color="transparent",border_width=2,command=self.on_leave,border_color="green",text_color="green")
        self.M_butn = cTk.CTkButton(master= self.frame_control, text="M",width=30,fg_color="transparent",border_width=2,command=self.on_enter,border_color="green",text_color="green")
        self.M_butn.pack(padx = 10, pady=5)
        self.button_save_tvare = cTk.CTkButton(master=self.frame_control, text="Save Tvari", command=self.button_s_tvare_event,width=120,fg_color="transparent",border_width=2)
        self.label_kamera_slider = cTk.CTkLabel(master = self.frame_control, width= 120,text="aktivacia kamery")
        self.slider_kamera = cTk.CTkSlider(master=self.frame_control, from_=0, to=100, command=self.slider_event_kamera,number_of_steps=1,width= 120)
        self.label_detekcia_slider = cTk.CTkLabel(master=self.frame_control,width= 120,text="aktivacia detekcie")
        self.slider_detekcia = cTk.CTkSlider(master= self.frame_control, from_=0, to=100, command=self.slider_event_detekcia,number_of_steps=1, width= 120)
        self.butn_dochadza_sheet= cTk.CTkButton(master=self.frame_control, text="Tabulka",width=120,fg_color="transparent",border_width=2,command=self.butn_dochadzka_sheet_event)
        self.entry_t_save = cTk.CTkEntry(master= self.frame_control,width=120, height=20,fg_color="transparent",border_width=2,state="normal",placeholder_text="1.zadaj meno 2.zmackni cudlik")

        # MAIN FRAME --> pre dni , back button a datumy
        self.frame_dni = cTk.CTkFrame(master= self.gui,width=830,height=100,border_width=2, border_color="green")
        # MAIN FRAME --> pre ludi, ktori su v datach tam sa budu skorlovat a bude napisane ze kedy dopsli, ak nedosli tak prazdne| alebo ak tento den este nebol
        self.frame_ludia = cTk.CTkScrollableFrame(master=self.gui, width=830, height=380,border_width=2, border_color="green",label_anchor="w")
        # back button
        self.back_butn = cTk.CTkButton(master=self.frame_dni, text="Back", command=self.back_butn,width=95,height=40,fg_color="transparent",border_width=2,border_color="green",text_color="green")
        # label pod ktorou budu datumy, tu su dni ktore knim patria
        self.label_text_dni = cTk.CTkLabel(master=self.frame_dni,text="|   Monday   |   Tuesday   |   Wednesday   |   Thursday   |   Friday   |",width=700,height=40,text_color="green",font=("Arial",23))
        self.label_date_dni = cTk.CTkLabel(master=self.frame_dni,text=self.dni_str,width=700,height=40,text_color="green",font=("Arial",18.5))
        self.label_date = cTk.CTkLabel(master=self.frame_dni,text="holder",text_color="green",font=("Arial",13))
        self.label_ludia = cTk.CTkLabel(master=self.frame_ludia,text=self.return_ludia(),text_color="green",font=("Arial",13))
        self.label_monday = cTk.CTkLabel(master=self.frame_ludia,text=self.return_zoznam(self.mondays),text_color="red",font=("Arial",11))
        self.label_tuesday = cTk.CTkLabel(master=self.frame_ludia,text=self.return_zoznam(self.tuesdays),text_color="yellow",font=("Arial",11))
        self.label_wednesday = cTk.CTkLabel(master=self.frame_ludia,text=self.return_zoznam(self.wednesday),text_color="purple",font=("Arial",11))
        self.label_thursday = cTk.CTkLabel(master=self.frame_ludia,text=self.return_zoznam(self.thursday),text_color="blue",font=("Arial",11))
        self.label_friday = cTk.CTkLabel(master=self.frame_ludia,text=self.return_zoznam(self.frydays),text_color="white",font=("Arial",11))
        
    def return_ludia(self):
        names_text = "\n".join(self.all_the_people.keys())
        print(f"Adding names:\n{names_text}")
        return names_text

    def return_zoznam(self,zoznam):
        zoznam_text =  "\n".join(zoznam)
        print(f"Adding names:\n{zoznam_text}")
        return zoznam_text

    def update_inf(self):
        gui_c.label_ludia.configure(text=gui_c.return_ludia())
        gui_c.label_monday.configure(text=gui_c.return_zoznam(gui_c.mondays))
        gui_c.label_tuesday.configure(text=gui_c.return_zoznam(gui_c.tuesdays))
        gui_c.label_wednesday.configure(text=gui_c.return_zoznam(gui_c.wednesday))
        gui_c.label_thursday.configure(text=gui_c.return_zoznam(gui_c.thursday))
        gui_c.label_friday.configure(text=gui_c.return_zoznam(gui_c.frydays)) 
    
    def slider_event_kamera(self,value):
        self.KAMERA_SLIDER_VALUE = int(value)
        
    def slider_event_detekcia(self,value):
        self.DETEKCIA_SLIDER_VALUE = int(value)
        
    def button_s_tvare_event(self):
        if self.DETEKCIA_SLIDER_VALUE > 0 :
            meno = self.entry_t_save.get()
            face_o.ziskaj_tvar(face_o.faces,meno)
            
    def butn_dochadzka_sheet_event(self):
        # vypnutie predoslich veci nemalo by zastavit check dochadzky len to nebude vidno
        self.frame_control.place_forget()
        self.MENU_butn.pack_forget()
        self.label_kamera_slider.pack_forget()
        self.slider_kamera.pack_forget()
        self.label_detekcia_slider.pack_forget()
        self.slider_detekcia.pack_forget()
        self.button_save_tvare.pack_forget()
        self.entry_t_save.pack_forget()
        self.butn_dochadza_sheet.pack_forget() 
        self.frame_kamera.place_forget()
        self.label_kamera.pack_forget()
        # zapnutie podrebnych veci
        self.frame_dni.pack(pady=10,padx=10,anchor="n")
        self.frame_dni.pack_propagate(False)
        self.frame_dni.columnconfigure(0, weight=1)  
        self.frame_dni.columnconfigure(1, weight=0)
        self.frame_ludia.pack(pady=10,padx=10,anchor="s")
        self.back_butn.place(relx=0.03, rely=0.08, anchor="nw")
        self.label_text_dni.place(relx=0.99, rely=0.05, anchor="ne")
        self.label_date_dni.place(relx=0.99, rely=0.9, anchor="se")
        self.label_date.place(relx=0.01,rely=0.9,anchor="sw")
        self.label_ludia.grid(row=0, column=0, padx=(0, 0))
        self.label_monday.grid(row=0, column=1, padx=(2, 2))
        self.label_tuesday.grid(row=0, column=2, padx=(2, 2))
        self.label_wednesday.grid(row=0, column=3, padx=(2, 2))
        self.label_thursday.grid(row=0, column=4, padx=(2, 2))
        self.label_friday.grid(row=0, column=5, padx=(2, 2))
        self.frame_ludia.columnconfigure(0, weight=1)
        self.frame_ludia.columnconfigure(1, weight=2)
        self.frame_ludia.columnconfigure(2, weight=2)
        self.frame_ludia.columnconfigure(3, weight=2)
        self.frame_ludia.columnconfigure(4, weight=2)
        self.frame_ludia.columnconfigure(5, weight=2)
        
    def back_butn(self):
        # vypnutie
        self.frame_dni.pack_forget()
        self.frame_ludia.pack_forget()
        self.back_butn.place_forget()
        self.label_text_dni.place_forget()
        self.label_date.place_forget()
        self.label_date_dni.place_forget()
        self.label_ludia.place_forget()
        # zapnutie - lavy panel
        self.frame_control.place(relx=0, rely=0, anchor="nw",relheight=1.0)
        self.frame_control.configure(width=150)
        self.MENU_butn.pack(padx=10,pady=5)
        self.label_kamera_slider.pack(pady=15, padx=5)
        self.slider_kamera.pack()
        self.label_detekcia_slider.pack(pady=15, padx=5)
        self.slider_detekcia.pack()
        self.button_save_tvare.pack(pady=(25, 15), padx=5)
        self.entry_t_save.pack(pady=15, padx=5)
        #entry_t_save.configure(state="disabled")
        self.butn_dochadza_sheet.pack(pady=15, padx=5)
        # zapnutie kamera
        self.frame_kamera.place(relx=1, rely=0, anchor="ne")
        self.label_kamera.pack(pady=15, padx=15)
        
    def on_enter(self):
        self.frame_control.configure(width=150)
        self.M_butn.pack_forget()
        self.MENU_butn.pack(padx=10,pady=5)
        self.label_kamera_slider.pack(pady=15, padx=5)
        self.slider_kamera.pack()
        self.label_detekcia_slider.pack(pady=15, padx=5)
        self.slider_detekcia.pack()
        self.button_save_tvare.pack(pady=(25, 15), padx=5)
        self.entry_t_save.pack(pady=15, padx=5)
        self.butn_dochadza_sheet.pack(pady=15, padx=5)
        
    def on_leave(self):
        self.frame_control.configure(width=75)
        self.MENU_butn.pack_forget()
        self.M_butn.pack(padx=10,pady=5)
        self.label_kamera_slider.pack_forget()
        self.slider_kamera.pack_forget()
        self.label_detekcia_slider.pack_forget()
        self.slider_detekcia.pack_forget()
        self.button_save_tvare.pack_forget()
        self.entry_t_save.pack_forget()
        self.butn_dochadza_sheet.pack_forget() 
        
    
    def update_label_date(self):
        terajsi_cas = dt.now()
        formatovany_cas = terajsi_cas.strftime("%Y-%m-%d %H:%M:%S")
        self.label_date.configure(text=formatovany_cas)
        self.gui.after(1000, self.update_label_date) 

                   
class Excel:
    def __init__(self):
        self.EXCEL_FILE_PATH = "DATA/EXCEL/Socka_Data.xlsx"
        self.SHEET_NAME = 'Data'
        # nacitanie excel filu
        self.workbook = openpyxl.load_workbook(self.EXCEL_FILE_PATH)
        self.sheet = self.workbook[self.SHEET_NAME]
        # excel pre oznacenie columns pozuziva pismena ak ak tam date cisla tak to nefunguje takze tak no
        self.column_letters = [  "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z",
                            "AA", "AB", "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN", "AO", "AP", "AQ", "AR", "AS", "AT", "AU", "AV", "AW", "AX", "AY", "AZ",
                            "BA", "BB", "BC", "BD", "BE", "BF", "BG", "BH", "BI", "BJ", "BK", "BL", "BM", "BN", "BO", "BP", "BQ", "BR", "BS", "BT", "BU", "BV", "BW", "BX", "BY", "BZ",
                            "CA", "CB", "CC", "CD", "CE", "CF", "CG", "CH", "CI", "CJ", "CK", "CL", "CM", "CN", "CO", "CP", "CQ", "CR", "CS", "CT", "CU", "CV", "CW", "CX", "CY", "CZ",]
        # oznacenie stplpca, bude vzdy A lebo ukladame len mena ludi plus ich fotky
        self.column_letter = 'A'
        # vytvorenie cerveneho a zeleneho borderu pre zapis dochadzky
        self.red_border_color = colors.Color('ff0000') # cervena farba pre meskanie
        self.red_border = Border(left=Side(border_style="thin", color=self.red_border_color),
                        right=Side(border_style="thin", color=self.red_border_color),
                        top=Side(border_style="thin", color=self.red_border_color),
                        bottom=Side(border_style="thin", color=self.red_border_color))
        self.green_border_color = colors.Color('00ff00')  # zelena ak si prisiel na cas  P.S toto bude vidiet lne v exceli lebo som lenivy
        self.green_border = Border(left=Side(border_style="thin", color=self.green_border_color),
                        right=Side(border_style="thin", color=self.green_border_color),
                        top=Side(border_style="thin", color=self.green_border_color),
                        bottom=Side(border_style="thin", color=self.green_border_color))
        self.all_the_people = {}
        self.week_dates = []
        self.colls = []
        self.mondays = []
        self.tuesdays = []
        self.wednesday = []
        self.thursday = []
        self.frydays = []  

        
        self.datum_tyzdna()
        self.get_the_names()
        self.get_week_days_column()
        self.get_the_final_value()
        self.vytvor_datum_dna_v_d()
    
    
    def new_value(self,row_num,value):
        print("this is how, you remind me , of what i really am")
        current_datetime = dt.now()
        den = current_datetime.weekday() # 0 -> pondelok, 1 -> utorok, a tak dalej , nechce sa mi to pisat 
        value1 = value.strftime('%Y-%m-%d %H:%M:%S')
        if row_num -2 <= len(gui_c.all_the_people.keys()) :
            if den == 0:
                gui_c.mondays[row_num-2] = value1
            if den == 1:
                gui_c.tuesdays[row_num-2] = value1
            if den == 2:
                gui_c.wednesday[row_num-2] = value1
            if den == 3:
                gui_c.thursday[row_num-2] = value1
            if den == 4:
                gui_c.frydays[row_num-2] = value1
            else:
                pass
        elif row_num-2 > len(gui_c.all_the_people.keys()) :
            if den == 0:
                gui_c.mondays.append(value1)
            if den == 1:
                gui_c.tuesdays.append(value1)
            if den == 2:
                gui_c.wednesday.append(value1)
            if den == 3:
                gui_c.thursday.append(value1)
            if den == 4:
                gui_c.frydays.appendt(value1)
        print(gui_c.mondays, gui_c.tuesdays ,gui_c.wednesday, gui_c.thursday ,gui_c.frydays)
        
        
    def new_person(self,meno,row):
        gui_c.all_the_people[meno] = row
        gui_c.mondays.append('------------------')
        gui_c.tuesdays.append('------------------')
        gui_c.wednesday.append('------------------')
        gui_c.thursday.append('------------------')
        gui_c.frydays.append('------------------')

       
        
    def vytvor_datum_dna_v_d(self):
         #tato premena odstranila jeden cely forloop takze velmi nice, sluzi aby funckia na ulozenie dochadzky vedele v ktorom column sa nachadzame
        cas = dt.now()
        cas_datum_dnes = cas.replace(hour=0, minute=0, second=0, microsecond=0)
        row_num = 1
        for i in self.column_letters :
            cell_value =  self.sheet[f'{i}{row_num}'].value
            if cell_value == None and cell_value != cas_datum_dnes:
                self.sheet[f'{i}{1}'] = cas_datum_dnes
                self.dnesny_column = i
                self.workbook.save(self.EXCEL_FILE_PATH) 
                break
            elif cell_value == cas_datum_dnes:
                self.dnesny_column = i
                break
            else:
                pass
            
    def ulozenie_mena(self,meno):
        row_num = 1
        while True :
            row_num += 1
            print(row_num)
            cell_value =  self.sheet[f'{self.column_letter}{row_num}'].value
            if cell_value is None :
                self.sheet[f'{self.column_letter}{row_num}'] = meno
                self.new_person(meno=meno,row=row_num)
                gui_c.update_inf()
                self.workbook.save(self.EXCEL_FILE_PATH)
                break
            else:
                pass    
            
    def najdi_cell_row_collum(self,meno):
        for row in self.sheet.iter_rows(min_row=2, max_col=1, max_row=self.sheet.max_row):
            for cell in row:
                if cell.value == meno:
                    row_number = cell.row
                    return row_number
        else:
            return "nepoznam"  # vrati nepoznam, ak meno nie je v exceli
    
    def uloz_dochadzku(self,meno):
        row_num = self.najdi_cell_row_collum(meno)
        if row_num == "nepoznam" :
            pass
        else:
            cas = dt.now()
            cas_8_rano = cas.replace(hour=8, minute=0, second=0, microsecond=0)
            cell_value =  self.sheet[f'{self.dnesny_column}{row_num}'].value
            
            
            if cell_value == None :
                if cas >= cas_8_rano :
                    self.sheet[f'{self.dnesny_column}{row_num}'] = cas
                    cell = self.sheet[f'{self.dnesny_column}{row_num}']
                    self.workbook.save(self.EXCEL_FILE_PATH)
                    cell.border = self.red_border
                    self.new_value(row_num,cas)
                    gui_c.update_inf()
                   

                else:
                    self.sheet[f'{self.dnesny_column}{row_num}'] = cas
                    cell = self.sheet[f'{self.dnesny_column}{row_num}']
                    cell.border = self.green_border
                    self.workbook.save(self.EXCEL_FILE_PATH)
                    self.new_value(row_num,cas)
                    gui_c.update_inf()
    

            else:
                pass    
            
    def datum_tyzdna(self):
        today = dt.now()
        start_of_week = today - timedelta(days=today.weekday())  
        self.week_dates = [start_of_week + timedelta(days=i) for i in range(5)]  
        return f'|  {self.week_dates[0].strftime("%Y-%m-%d")}   |   {self.week_dates[1].strftime("%Y-%m-%d")}   |      {self.week_dates[2].strftime("%Y-%m-%d")}       |    {self.week_dates[3].strftime("%Y-%m-%d")}   | {self.week_dates[4].strftime("%Y-%m-%d")} |'

    def get_the_names(self):
        for row in self.sheet.iter_rows(min_row=2, max_col=1, max_row=self.sheet.max_row):
            for cell in row:
                #print(f'cell value {cell.value} a cell row {row}')
                if cell.value != None and cell.value.strip():
                    self.all_the_people[cell.value.strip()] = cell.row
                else:
                    break
                
    def get_week_days_column(self):
        for day in self.week_dates:
            day = day.replace(hour=0, minute=0, second=0, microsecond=0)
            print(day)
            found_match = False

            for col in self.sheet.iter_rows(min_col=2, max_row=1, max_col=self.sheet.max_column):
                #print(col)
                for cell in col:
                    if cell.value is not None and cell.value == day:
                        #print("ide to ideeeee",cell.value)
                        #colls[day] = column_letters[cell.column-2]
                        self.colls.append(self.column_letters[cell.column-2])
                        found_match = True
                        break 

                if found_match:
                    break  
            if not found_match:
                #colls[day] = 'Nein'
                self.colls.append("")
                
    def get_the_final_value(self):
        den = 0
        for column in self.colls :
            den += 1
            #print(ahhhhhh)
            for row in self.all_the_people.values() :
                if column != "" :
                    hod = self.sheet[f'{column}{row}'].value
                    #print(f'cislo: {ahhhhhh} hod: {hod}')
                    if hod != None :
                        hod = hod.strftime("%Y %m %d %H %M %S")
                        if den == 1 :
                            self.mondays.append(hod)
                        elif den == 2 :
                            self.tuesdays.append(hod)
                        elif den == 3 :
                            self.wednesday.append(hod)
                        elif den == 4 :
                            self.thursday.append(hod)
                        elif den == 5 :
                            self.frydays.append(hod)
                    else:
                        if den == 1 :
                            self.mondays.append('------------------') 
                        elif den == 2 :
                            self.tuesdays.append('------------------')
                        elif den == 3 :
                            self.wednesday.append('------------------')
                        elif den == 4 :
                            self.thursday.append('------------------')
                        elif den == 5 :
                            self.frydays.append('------------------')  
                else:
                    if den == 1 :
                        self.mondays.append('------------------') 
                    elif den == 2 :
                        self.tuesdays.append('------------------')
                    elif den == 3 :
                        self.wednesday.append('------------------')
                    elif den == 4 :
                        self.thursday.append('------------------')
                    elif den == 5 :
                        self.frydays.append('------------------') 
                        
                        
face_o = FaceFunctions()
excel = Excel()
gui_c = CustomTkinter(gui,m=excel.mondays,tu=excel.tuesdays,w=excel.wednesday,th=excel.thursday,f=excel.frydays,people=excel.all_the_people)